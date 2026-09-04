"""Rebuild wages.json from a BLS OEWS national file.

There was no script for this. wages.json was built by hand for the May 2024
release, which is exactly why the data quietly fell a release behind while a
scheduled job watched O*NET for staleness and nothing watched BLS.

    python3 scripts/data/build_wages.py 25          # May 2025 release
    python3 scripts/data/build_wages.py 24 --check  # reproduce the old file

--check rebuilds an existing release and diffs against wages.json. Run it
before trusting a new release: if the logic cannot reproduce what is already
in the org, it should not be used to replace it.

BLS returns 403 to a default User-Agent, so the download must carry contact
details. See scripts/data/README.md.

Join: our ONET_Code__c minus its suffix -> BLS OCC_CODE. Where BLS publishes no
detailed row, fall back to the broad group (SOC ending 0) and FLAG it, because
a broad-group figure presented as this exact role's pay is a wrong number with
a federal source attached.
"""
import json, os, subprocess, sys, zipfile
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ORG  = os.environ.get("NM_ORG", "dreamforce-hackathon")
UA   = "NextMission/1.0 (andrew.purpura@jaxconsult.com)"


def occupations():
    """The 1,016 O*NET occupations we actually ground on, from the org."""
    out = subprocess.run(
        ["sf", "data", "query", "--target-org", ORG, "--json", "-q",
         "SELECT Name, ONET_Code__c FROM NM_Occupation__c"],
        capture_output=True, text=True, timeout=600)
    recs = json.loads(out.stdout)["result"]["records"]
    return [(r["Name"], r["ONET_Code__c"]) for r in recs if r.get("ONET_Code__c")]


def fetch(year, workdir):
    zip_path = os.path.join(workdir, "oesm%snat.zip" % year)
    if not os.path.exists(zip_path):
        url = "https://www.bls.gov/oes/special-requests/oesm%snat.zip" % year
        print("downloading %s" % url)
        subprocess.run(["curl", "-sS", "-A", UA, "-o", zip_path, url, "--max-time", "300"],
                       check=True, timeout=400)
    with zipfile.ZipFile(zip_path) as z:
        name = [n for n in z.namelist() if n.endswith(".xlsx")][0]
        z.extract(name, workdir)
        return os.path.join(workdir, name)


def read_oews(xlsx):
    """OCC_CODE -> row, for every code BLS publishes at any level."""
    wb = openpyxl.load_workbook(xlsx, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    hdr = next(rows)
    col = {h: i for i, h in enumerate(hdr) if h}

    def num(v):
        # BLS uses '*' for suppressed and '#' for "at or above $239,200".
        if v is None or v in ("*", "**", "#"):
            return None
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return None

    table = {}
    for r in rows:
        code = r[col["OCC_CODE"]]
        if not code:
            continue
        table[code] = {
            "title":  r[col["OCC_TITLE"]],
            "group":  r[col["O_GROUP"]],
            "median": num(r[col["A_MEDIAN"]]),
            "p10":    num(r[col["A_PCT10"]]),
            "p90":    num(r[col["A_PCT90"]]),
            "emp":    num(r[col["TOT_EMP"]]),
        }
    return table


def build(year, workdir):
    table = read_oews(fetch(year, workdir))
    occs  = occupations()

    out, detailed, broad, missing = [], 0, 0, 0
    for name, onet in sorted(occs, key=lambda x: x[1]):
        soc = onet.split(".")[0]
        row = table.get(soc)
        level = "detailed"
        if row is None or row["median"] is None:
            # BLS publishes many roles only at the broad group, whose SOC ends 0.
            row = table.get(soc[:6] + "0")
            level = "broad"
        if row is None or row["median"] is None:
            missing += 1
            continue
        if level == "detailed":
            detailed += 1
        else:
            broad += 1
        out.append({
            "onet": onet, "soc": soc, "level": level, "title": row["title"],
            "median": row["median"], "p10": row["p10"], "p90": row["p90"],
            "emp": row["emp"],
        })

    print("\n  occupations:        %d" % len(occs))
    print("  detailed SOC match: %d" % detailed)
    print("  broad group only:   %d  (must be disclosed by the agent)" % broad)
    print("  no published median:%d" % missing)
    print("  coverage:           %.1f%%" % (100.0 * len(out) / len(occs)))
    return out


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(2)
    year  = sys.argv[1]
    check = "--check" in sys.argv
    workdir = os.environ.get("NM_WORKDIR", "/tmp")

    rows = build(year, workdir)
    target = os.path.join(HERE, "wages.json")

    if check:
        existing = json.load(open(target))
        a = {r["onet"]: (r["median"], r["p10"], r["p90"], r["emp"]) for r in existing}
        b = {r["onet"]: (r["median"], r["p10"], r["p90"], r["emp"]) for r in rows}
        only_a = sorted(set(a) - set(b))
        only_b = sorted(set(b) - set(a))
        diff   = sorted(k for k in set(a) & set(b) if a[k] != b[k])
        print("\n--check against the committed wages.json")
        print("  in file, not rebuilt: %d" % len(only_a))
        print("  rebuilt, not in file: %d" % len(only_b))
        print("  same code, different figures: %d" % len(diff))
        for k in diff[:5]:
            print("    %s  file=%s  rebuilt=%s" % (k, a[k], b[k]))
        ok = not only_a and not only_b and not diff
        print("\n%s" % ("REBUILD REPRODUCES THE COMMITTED FILE EXACTLY"
                        if ok else "REBUILD DOES NOT MATCH. Do not use it to replace data."))
        sys.exit(0 if ok else 1)

    with open(target, "w") as fh:
        json.dump(rows, fh)
    print("\nwrote %s (%d rows)" % (target, len(rows)))


if __name__ == "__main__":
    main()
